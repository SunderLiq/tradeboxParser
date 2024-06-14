import requests
from bs4 import BeautifulSoup
from lxml import etree 
import pandas as pd
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# URL главной страницы
base_url = "http://tradebox.shop"

# Функция для извлечения ссылок на категории товаров
def get_category_links():
    response = requests.get(base_url)
    soup = BeautifulSoup(response.text, 'html.parser')
    # Извлекаем ссылки на категории
    category_links = [base_url + a['href'] for a in soup.select("div.header__catalog-dropdown.clear > ul > li > a")]
    return category_links

# Функция для извлечения данных о товарах из одной категории
def get_products_from_category(category_url):
    products = []
    while category_url:
        response = requests.get(category_url)
        soup = BeautifulSoup(response.text, 'html.parser')
        dom = etree.HTML(str(soup))
        category = (dom.xpath('/html/body/main/section/div/div/div[1]/div/ul/li[3]')[0].text) 
        
        # Извлекаем данные о товарах
        for product in soup.select("div.card"):
            title = product.select_one("a.card__title").get_text(strip=True) 
            product_url = product.select_one("a.card__title").get('href')
            print(product_url)
            price = product.select_one("div.card__price").get_text(strip=True)
            if (price == "") :
                price = "Цена не указана"
            products.append({
                'Название': title,
                'Категория' : category,
                'Цена': price,
                'Ссылка' : product_url                
            })
        
        # Проверка наличия следующей страницы
        next_page = soup.select_one("a.next-page")
        category_url = base_url + next_page['href'] if next_page else None
    
    return products

# Функция для сбора всех данных с сайта
def scrape_all_products():
    all_products = []
    
    # Получаем ссылки на все категории
    category_links = get_category_links()
    
    for category_url in category_links:
        print(f"Собираем данные с {category_url}")
        products = get_products_from_category(category_url)
        all_products.extend(products)
    
    return all_products

# Сохранение данных в .xlsx файл
def save_to_excel(data, filename):
    # Преобразуем список словарей в DataFrame
    df = pd.DataFrame(data)
    # Сохраняем DataFrame в .xlsx файл
    df.to_excel(filename, index=False)

# Запуск сбора данных и сохранение их в файл
all_products = scrape_all_products()
save_to_excel(all_products, 'products.xlsx')

print(f"Сохранено {len(all_products)} товаров в 'products.xlsx'")

# Сохранение данных в CSV-файл
def save_to_csv(data, filename):
    # Проверяем, что список не пуст и содержит словари
    if not data or not isinstance(data[0], dict):
        print("Нет данных для сохранения или неверный формат данных.")
        return
    
    # Определяем заголовки столбцов
    keys = data[0].keys()
    with open(filename, 'w', newline='', encoding='utf-8') as output_file:
        dict_writer = csv.DictWriter(output_file, fieldnames=keys)
        dict_writer.writeheader()
        dict_writer.writerows(data)

# Запуск сбора данных и сохранение их в файл
all_products = scrape_all_products()
save_to_csv(all_products, 'products.csv')

print(f"Сохранено {len(all_products)} товаров в 'products.csv'")

workbook = openpyxl.load_workbook('products.xlsx')
worksheet = workbook.active  # Выбор активного листа

# Установите ширину столбцов
column_widths = {
    'A': 350,  # Название
    'B': 250,  # Категория
    'C': 112,  # Цена
    'D': 520   # Ссылка
}

# Применяем ширину к каждому столбцу
for col, width in column_widths.items():
    worksheet.column_dimensions[col].width = width / 7.2  # Делим на 7.2 для приведения к более подходящему размеру в Excel

# Включаем перенос текста для всех ячеек в столбцах
for row in worksheet.iter_rows(min_row=1, max_col=4, max_row=worksheet.max_row):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

# Сохраняем изменения в файл
workbook.save('products.xlsx')